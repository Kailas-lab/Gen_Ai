import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import io

st.set_page_config(page_title="Excel Comparison Tool", layout="wide")

st.title("📊 R0 vs R1 Excel Comparison Tool")
st.write("Upload two Excel files and click **Run Comparison**")

# === File Upload ===
r0_file = st.file_uploader("Upload R0.xlsx", type=["xlsx"])
r1_file = st.file_uploader("Upload R1.xlsx", type=["xlsx"])

# === Run Button ===
run_btn = st.button("▶ Run Comparison")

if run_btn:

    if not r0_file or not r1_file:
        st.warning("⚠️ Please upload both R0.xlsx and R1.xlsx before running.")
        st.stop()

    try:
        # === Load Data ===
        r0_df = pd.read_excel(r0_file, dtype=str).fillna("")
        r1_df = pd.read_excel(r1_file, dtype=str).fillna("")

        # Ensure Tag column exists
        if 'Tag' not in r0_df.columns or 'Tag' not in r1_df.columns:
            st.error("❌ Both files must contain a 'Tag' column")
            st.stop()

        # Drop duplicates
        r0_df = r0_df.drop_duplicates(subset='Tag').set_index('Tag')
        r1_df = r1_df.drop_duplicates(subset='Tag').set_index('Tag')

        # Column handling
        r0_columns = list(r0_df.columns)
        all_columns = sorted(
            set(r0_df.columns).union(set(r1_df.columns)),
            key=lambda x: (r0_columns.index(x) if x in r0_columns else float('inf'))
        )

        all_tags = sorted(set(r0_df.index).union(set(r1_df.index)))
        comparison_rows = []

        # === Comparison Logic ===
        for tag in all_tags:

            if tag not in r0_df.index:
                row = {"Tag": tag, "Change_Type": "✅ Added in R1"}
                row.update({col: r1_df.loc[tag].get(col, "") for col in all_columns})
                row["Change_Summary"] = ""
                comparison_rows.append(row)

            elif tag not in r1_df.index:
                row = {"Tag": tag, "Change_Type": "❌ Removed in R1"}
                row.update({col: r0_df.loc[tag].get(col, "") for col in all_columns})
                row["Change_Summary"] = ""
                comparison_rows.append(row)

            else:
                row_r0 = r0_df.loc[tag]
                row_r1 = r1_df.loc[tag]
                row_data = {"Tag": tag}
                summary = []
                changes_exist = False

                for col in all_columns:
                    val_r0 = row_r0.get(col, "")
                    val_r1 = row_r1.get(col, "")
                    if str(val_r0).strip() != str(val_r1).strip():
                        row_data[col] = f"{val_r0} → {val_r1}"
                        summary.append(f"{col}: {val_r0} → {val_r1}")
                        changes_exist = True
                    else:
                        row_data[col] = val_r1

                row_data["Change_Type"] = "✏️ Modified" if changes_exist else "No Change"
                row_data["Change_Summary"] = " | ".join(summary)
                comparison_rows.append(row_data)

        # === Final DataFrame ===
        comparison_df = pd.DataFrame(comparison_rows)
        final_columns = ["Tag", "Change_Type"] + all_columns + ["Change_Summary"]
        comparison_df = comparison_df[final_columns]

        # === Preview ===
        st.subheader("🔍 Comparison Preview")
        st.dataframe(comparison_df, use_container_width=True)

        # === Excel Creation ===
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison Summary"
        highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for r_idx, row in enumerate(dataframe_to_rows(comparison_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, str) and "→" in value:
                    cell.fill = highlight

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%d_%m_%Y_%H_%M")
        filename = f"Vimal_Comparison_R0_vs_R1_{timestamp}.xlsx"

        st.download_button(
            "⬇️ Download Comparison Excel",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success("✅ Comparison completed successfully!")

    except Exception as e:
        st.error(f"⚠️ Error occurred: {e}")
